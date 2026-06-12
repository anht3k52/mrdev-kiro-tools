# -*- coding: utf-8 -*-
"""idc_browser_login.py — Tu dong login trang device-code AWS IAM Identity Center
bang Playwright (Chromium SACH: khong cache / cookie / profile).

Luong (da soi DOM that tu d-9066713dd7.awsapps.com):
    1. Username page  : input[type=text] (#awsui-input-0)        -> "Next"
    2. Password page  : input[type=password] (#awsui-input-1)    -> "Sign in"
    3. (LAN DAU) Doi mat khau: 2 o password (new + confirm)      -> "Confirm"
    4. Authorization requested (hien user_code)                  -> "Confirm and continue"
    5. Allow kiro-oauth-client to access your data?              -> "Allow access"
    -> device-code da APPROVE => device_code_auth.poll_for_token() lay token durable.

Tool tu PHAT HIEN co form doi mat khau hay khong (qua so o password hien),
nen xu ly duoc ca 2 dang: lan dau (can doi pass) va lan sau (khong can).
"""
from __future__ import annotations

import math
import random
import re
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Callable, List, Optional

from playwright.sync_api import sync_playwright, TimeoutError as PWTimeoutError


# Mac dinh
DEFAULT_IDC_START_URL = "https://d-9066713dd7.awsapps.com/start"
DEFAULT_NEW_PASSWORD = "Kiro@Durable2026#"   # du manh cho policy AWS (>=8, hoa/thuong/so/ky tu)

# Text marker (lowercase) de nhan dien trang thai
INCORRECT_MARKERS = (
    "incorrect", "is not correct", "could not sign", "couldn't sign",
    "authentication failed", "username or password", "try again later",
    "we couldn't", "no account", "does not exist",
)
MFA_MARKERS = (
    "multi-factor", "mfa device", "authenticator app", "verification code",
    "register mfa", "one-time passcode", "security key", "passkey",
)
SUCCESS_MARKERS = (
    "request approved", "approved", "you can close", "added to your devices",
    "you may close", "request was approved",
)


@dataclass
class LoginOutcome:
    ok: bool
    changed_password: bool = False
    error: str = ""


@dataclass
class AccountRow:
    idx: int                 # so dong (xlsx: row 1-based; txt: line 1-based)
    email: str
    password: str
    proxy: str = ""
    kind: str = "xlsx"       # "xlsx" | "txt"


# ======================================================================
# Doc / ghi file account
# ======================================================================
_SEP_RE = re.compile(r"[\t,;|]")


def read_accounts_file(path: str | Path) -> List[AccountRow]:
    """Doc account tu .xlsx (Email|Password|Proxy) hoac .txt/.csv (email:password)."""
    p = Path(path)
    ext = p.suffix.lower()
    if ext in (".xlsx", ".xlsm"):
        return _read_xlsx(p)
    return _read_txt(p)


def _read_xlsx(p: Path) -> List[AccountRow]:
    from openpyxl import load_workbook
    wb = load_workbook(p)
    ws = wb.active
    out: List[AccountRow] = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row:
            continue
        email = _cell(row[0]) if len(row) > 0 else ""
        if not email or email.lower() in ("email", "username", "user", "account"):
            continue
        pw = _cell(row[1]) if len(row) > 1 else ""
        proxy = _cell(row[2]) if len(row) > 2 else ""
        out.append(AccountRow(idx=idx, email=email, password=pw, proxy=proxy, kind="xlsx"))
    wb.close()
    return out


def _read_txt(p: Path) -> List[AccountRow]:
    out: List[AccountRow] = []
    for i, raw in enumerate(p.read_text(encoding="utf-8", errors="replace").splitlines(), start=1):
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        # BO comment ket qua tool tu ghi-back ("   # OK/LOGIN FAIL ...").
        # Dung pattern ">=2 khoang trang + #" -> KHONG dung cham vao '#' trong password
        # (vd "Kiro@Durable2026#" giu nguyen vi khong co khoang trang truoc '#').
        line = re.split(r"\s{2,}#", line, maxsplit=1)[0].rstrip()
        if not line:
            continue
        # tach theo tab/comma/semicolon/pipe truoc, fallback dau ':'
        parts = _SEP_RE.split(line)
        if len(parts) < 2:
            parts = line.split(":", 2)  # email:pass[:proxy] (email khong chua ':')
        parts = [x.strip() for x in parts]
        email = parts[0] if parts else ""
        pw = parts[1] if len(parts) > 1 else ""
        # chap nhan username thuan (khong can @); can co password de tranh dong rac
        if not email or not pw:
            continue
        if email.lower() in ("email", "username", "user", "account"):
            continue
        proxy = parts[2] if len(parts) > 2 else ""
        out.append(AccountRow(idx=i, email=email, password=pw, proxy=proxy, kind="txt"))
    return out


def write_account_result(
    path: str | Path,
    acc: AccountRow,
    new_password: Optional[str],
    result: str,
) -> None:
    """Ghi ket qua + (neu doi pass) cap nhat password moi vao file de lan sau dung dung."""
    p = Path(path)
    try:
        if acc.kind == "xlsx":
            from openpyxl import load_workbook
            wb = load_workbook(p)
            ws = wb.active
            if new_password:
                ws.cell(row=acc.idx, column=2, value=new_password)   # cot B = Password
            ws.cell(row=acc.idx, column=4, value=result)             # cot D = Result
            wb.save(p)
            wb.close()
        else:
            # txt: rewrite dung dong, cap nhat password neu doi
            lines = p.read_text(encoding="utf-8", errors="replace").splitlines()
            if 1 <= acc.idx <= len(lines):
                if new_password:
                    sep = ":"
                    m = _SEP_RE.search(lines[acc.idx - 1])
                    if m:
                        sep = m.group(0)
                    pieces = [acc.email, new_password]
                    if acc.proxy:
                        pieces.append(acc.proxy)
                    lines[acc.idx - 1] = sep.join(pieces) + f"    # {result}"
                else:
                    base = lines[acc.idx - 1].split("    #", 1)[0].rstrip()
                    lines[acc.idx - 1] = base + f"    # {result}"
                p.write_text("\n".join(lines) + "\n", encoding="utf-8")
    except Exception:
        pass  # ghi ket qua that bai -> khong chan luong chinh


def _cell(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


# ======================================================================
# Playwright helpers
# ======================================================================

def _body_text(page) -> str:
    try:
        return (page.evaluate("() => document.body ? document.body.innerText : ''") or "")
    except Exception:
        return ""


def _count_visible(page, selector: str) -> int:
    try:
        return page.locator(f"{selector} >> visible=true").count()
    except Exception:
        try:
            return page.locator(selector).count()
        except Exception:
            return 0


def _maybe_cookie_accept(page, log) -> None:
    """Dong banner cookie AWS neu dang chan."""
    for sel in (
        'button[aria-label="Accept all cookies"]',
        'button[data-id="awsccc-cb-btn-accept"]',
    ):
        try:
            b = page.locator(sel).first
            if b.count() and b.is_visible():
                b.click(timeout=1500)
                return
        except Exception:
            pass


def _click_button(page, texts, log, timeout=2500) -> bool:
    """Bam button hien thi co text khop (exact uu tien, fallback contains)."""
    for t in texts:
        makers = (
            page.get_by_role("button", name=t, exact=True),
            page.get_by_role("button", name=t),
            page.locator(f'button:has-text("{t}")'),
        )
        for mk in makers:
            try:
                b = mk.first
                if b.count() and b.is_visible():
                    b.click(timeout=timeout)
                    log(f"    -> click '{t}'")
                    return True
            except Exception:
                pass
    return False


def _fill_first_visible(page, selector: str, value: str, timeout=15000) -> bool:
    """Dien + VERIFY gia tri da vao field (chong race khi chay da luong)."""
    try:
        loc = page.locator(f"{selector} >> visible=true").first
        loc.wait_for(state="visible", timeout=timeout)
        loc.click(timeout=4000)
        try:
            loc.fill("")
        except Exception:
            pass
        loc.fill(value)
        # verify: cho chac field nhan dung gia tri truoc khi submit
        for _ in range(4):
            try:
                if (loc.input_value() or "") == value:
                    break
            except Exception:
                break
            try:
                loc.fill(value)
            except Exception:
                pass
            time.sleep(0.2)
        time.sleep(0.35)  # settle cho SPA dang ky onChange
        return True
    except Exception:
        return False


def _fill_nth_password(page, idx: int, value: str) -> None:
    """Dien o password thu idx (cho form doi mat khau), co verify."""
    try:
        loc = page.locator("input[type=password] >> visible=true").nth(idx)
        loc.click(timeout=3000)
        loc.fill(value)
        for _ in range(3):
            try:
                if (loc.input_value() or "") == value:
                    break
            except Exception:
                break
            loc.fill(value)
            time.sleep(0.15)
    except Exception:
        pass


def _visible_error_text(page) -> str:
    """Lay text loi THAT tu element alert (de log ly do that bai). '' neu khong co."""
    sel = ("[role=alert], [aria-live='assertive'], [aria-live='polite'], "
           "[class*='rror']")
    try:
        els = page.locator(sel)
        for i in range(min(els.count(), 8)):
            e = els.nth(i)
            try:
                if e.is_visible():
                    t = " ".join((e.inner_text() or "").split())
                    if len(t) >= 4:
                        return t[:140]
            except Exception:
                pass
    except Exception:
        pass
    return ""


def _wait_metadata1(page, timeout: float = 30.0, min_len: int = 500) -> bool:
    """Cho field an metadata1 (token chong-bot cua AWS, do JS sinh sau ~1.5s+) day du
    truoc khi submit. Quan trong khi chay DA LUONG (CPU cham -> JS sinh token lau hon).
    Tra True khi san sang (hoac khong co field do). False neu het timeout."""
    end = time.time() + timeout
    while time.time() < end:
        try:
            ln = page.evaluate(
                "() => { const el = document.querySelector('input[name=metadata1]');"
                " return el ? (el.value || '').length : -1; }")
        except Exception:
            ln = -1
        if ln == -1:      # khong co field -> khong can cho
            return True
        if ln >= min_len:
            return True
        time.sleep(0.4)
    return False


def _wait_password_resolve(page, timeout: float = 12.0) -> str:
    """Cho trang password chuyen tiep. TIN HIEU CAU TRUC (khong dua vao text body):
    'gone'  = da roi trang password (so o password != 1, hoac hien nut consent) -> OK,
    'stuck' = sau timeout van ket o trang password (sai pass / chua submit).
    """
    end = time.time() + timeout
    while time.time() < end:
        if _count_visible(page, "input[type=password]") != 1:
            return "gone"   # 0 = qua consent; >=2 = sang form doi mat khau
        if (_is_button_visible(page, "Allow access")
                or _is_button_visible(page, "Confirm and continue")):
            return "gone"
        time.sleep(0.4)
    return "stuck"


# ======================================================================
# Drive login
# ======================================================================

def _tile_args(window_index: int, window_count: int,
               screen_w: int, screen_h: int) -> list:
    """Tinh --window-position/--window-size de xep browser thanh luoi (headed)."""
    if window_count <= 1:
        return ["--start-maximized"]
    cols = min(4, max(1, math.ceil(math.sqrt(window_count))))
    rows = max(1, math.ceil(window_count / cols))
    w = max(560, screen_w // cols)
    h = max(520, screen_h // rows)
    slot = window_index % (cols * rows)
    x = (slot % cols) * w
    y = (slot // cols) * h
    return [f"--window-position={x},{y}", f"--window-size={w},{h}"]


def drive_login(
    verification_uri_complete: str,
    email: str,
    password: str,
    new_password: str = DEFAULT_NEW_PASSWORD,
    log: Callable[[str], None] = print,
    headless: bool = False,
    stop_event=None,
    proxy: str = "",
    timeout_s: int = 220,
    window_index: int = 0,
    window_count: int = 1,
    screen_w: int = 1920,
    screen_h: int = 1040,
    debug_dir: str = "",
) -> LoginOutcome:
    """Mo Chromium sach, tu login trang device-code den khi bam 'Allow access'."""
    launch_args = ["--disable-blink-features=AutomationControlled"]
    if not headless:
        launch_args += _tile_args(window_index, window_count, screen_w, screen_h)
    launch_kwargs = {"headless": headless, "args": launch_args}
    if proxy:
        srv = proxy if "://" in proxy else f"http://{proxy}"
        launch_kwargs["proxy"] = {"server": srv}

    with sync_playwright() as pw:
        browser = pw.chromium.launch(**launch_kwargs)
        ctx = browser.new_context(no_viewport=not headless)
        page = ctx.new_page()
        page.set_default_timeout(20000)
        try:
            return _flow(page, verification_uri_complete, email, password,
                         new_password, log, stop_event, timeout_s, debug_dir)
        except Exception as e:
            return LoginOutcome(False, error=f"loi browser: {e}")
        finally:
            try:
                browser.close()
            except Exception:
                pass


def _flow(page, url, email, password, new_password, log, stop_event, timeout_s,
          debug_dir: str = "") -> LoginOutcome:
    safe = "".join(c if c.isalnum() else "_" for c in (email or "acc"))[:20]
    shot = {"i": 0}

    def _shot(tag: str):
        if not debug_dir:
            return
        try:
            from pathlib import Path as _P
            _P(debug_dir).mkdir(exist_ok=True, parents=True)
            page.screenshot(path=str(_P(debug_dir) / f"{safe}_{shot['i']:02d}_{tag}.png"))
        except Exception:
            pass
        shot["i"] += 1

    # jitter nho de cac luong khong dong bo (giam tranh chap khi chay song song)
    time.sleep(random.uniform(0.1, 1.6))
    log("    mo trang login (clean, no cache)...")
    page.goto(url, wait_until="domcontentloaded")
    try:
        page.wait_for_selector("input", timeout=25000)
    except Exception:
        pass

    deadline = time.time() + timeout_s
    changed = False
    pw_attempts = 0
    user_attempts = 0
    idle_rounds = 0
    seen_password = False

    while time.time() < deadline:
        if stop_event is not None and stop_event.is_set():
            return LoginOutcome(False, changed, "Da huy")

        _maybe_cookie_accept(page, log)
        body = _body_text(page).lower()

        # --- MFA (user da xac nhan khong co; neu gap -> bao loi ro) ---
        if any(m in body for m in MFA_MARKERS):
            return LoginOutcome(False, changed,
                                "Gap buoc MFA/2FA - account nay co bao mat 2 lop, dung lai.")

        # --- da Allow xong / approved ---
        if any(m in body for m in SUCCESS_MARKERS):
            log("    -> approved")
            return LoginOutcome(True, changed)

        pw_n = _count_visible(page, "input[type=password]")
        text_n = _count_visible(page, "input[type=text]:not([type=hidden])")
        has_allow = _is_button_visible(page, "Allow access")
        has_confirm = _is_button_visible(page, "Confirm and continue")

        # --- consent: Allow access (cuoi cung) ---
        if has_allow:
            if _click_button(page, ["Allow access"], log):
                time.sleep(2.0)
                return LoginOutcome(True, changed)   # approve xong -> poll se lay token

        # --- consent: Confirm and continue (xac nhan user_code) ---
        if has_confirm:
            _click_button(page, ["Confirm and continue"], log)
            idle_rounds = 0
            time.sleep(2.0)
            continue

        # --- DOI MAT KHAU lan dau: >=2 o password ---
        if pw_n >= 2:
            log("    form DOI MAT KHAU lan dau -> dat password moi")
            if pw_n >= 3:
                # current / new / confirm
                _fill_nth_password(page, 0, password)
                _fill_nth_password(page, 1, new_password)
                _fill_nth_password(page, 2, new_password)
            else:
                _fill_nth_password(page, 0, new_password)
                _fill_nth_password(page, 1, new_password)
            time.sleep(0.3)
            _wait_metadata1(page)
            _click_button(
                page,
                ["Confirm", "Change password", "Update password", "Set new password",
                 "Set password", "Save changes", "Save", "Continue", "Submit"],
                log,
            )
            changed = True
            idle_rounds = 0
            time.sleep(2.5)
            continue

        # --- PASSWORD page: 1 o password ---
        if pw_n == 1:
            seen_password = True
            pw_attempts += 1
            log(f"    nhap password (lan {pw_attempts})")
            _wait_metadata1(page)            # cho token anti-bot truoc khi dien
            _fill_first_visible(page, "input[type=password]", password)
            _shot("pw_filled")
            if debug_dir:
                try:
                    val = page.locator("input[type=password] >> visible=true").first.input_value()
                    log(f"    [debug] o password dang chua: {val!r} (khop pass={val == password})")
                except Exception:
                    pass
            if not _wait_metadata1(page):
                log("    WARN: metadata1 chua san sang (van thu submit)")
            if not _click_button(page, ["Sign in", "Continue", "Next"], log):
                try:
                    page.keyboard.press("Enter")
                except Exception:
                    pass
            # roi trang password = OK; ket lai = sai pass / chua submit
            st = _wait_password_resolve(page, timeout=12.0)
            _shot(f"after_signin_{st}")
            if st == "gone":
                idle_rounds = 0
                continue
            detail = _visible_error_text(page)
            if pw_attempts >= 2:
                return LoginOutcome(
                    False, changed,
                    f"Login that bai o trang password: {detail or 'khong qua duoc (sai pass?)'}")
            log(f"    chua qua password ({detail or 'thu lai'}) -> thu lai")
            idle_rounds = 0
            continue

        # --- USERNAME page: co text input, chua co password ---
        if text_n >= 1 and pw_n == 0 and not has_allow and not has_confirm:
            # Da qua password roi ma quay lai username = sign-in bi tu choi/reset
            if seen_password:
                detail = _visible_error_text(page)
                _shot("reset_to_username")
                return LoginOutcome(
                    False, changed,
                    f"Sign-in bi tu choi (reset ve username): {detail or 'anti-bot/sai pass'}")
            user_attempts += 1
            if user_attempts > 3:
                return LoginOutcome(False, changed, "Khong qua duoc buoc username (email sai?).")
            log("    nhap username/email")
            _wait_metadata1(page)
            _fill_first_visible(page, "#awsui-input-0, input[type=text]", email)
            _shot("user_filled")
            if not _wait_metadata1(page):
                log("    WARN: metadata1 chua san sang (van thu submit)")
            if not _click_button(page, ["Next", "Continue", "Sign in"], log):
                try:
                    page.keyboard.press("Enter")
                except Exception:
                    pass
            # cho chuyen sang trang password (tranh re-fire username khi parallel cham)
            try:
                page.wait_for_selector("input[type=password] >> visible=true", timeout=12000)
            except Exception:
                pass
            idle_rounds = 0
            continue

        # --- khong co gi de lam ---
        idle_rounds += 1
        # da qua password ma khong con form/nut -> coi nhu approve xong
        if pw_attempts >= 1 and idle_rounds >= 4:
            log("    khong con form/nut consent -> coi nhu da xong")
            return LoginOutcome(True, changed)
        time.sleep(1.2)

    return LoginOutcome(False, changed, "Timeout - khong hoan tat login trong thoi gian cho.")


def _is_button_visible(page, text: str) -> bool:
    try:
        b = page.get_by_role("button", name=text, exact=True).first
        if b.count() and b.is_visible():
            return True
    except Exception:
        pass
    try:
        b = page.locator(f'button:has-text("{text}")').first
        return bool(b.count() and b.is_visible())
    except Exception:
        return False


__all__ = [
    "LoginOutcome", "AccountRow",
    "read_accounts_file", "write_account_result",
    "drive_login", "DEFAULT_IDC_START_URL", "DEFAULT_NEW_PASSWORD",
]
